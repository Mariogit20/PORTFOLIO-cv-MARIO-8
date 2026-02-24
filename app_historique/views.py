from __future__ import annotations

import csv
import json
from datetime import timedelta
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_GET, require_POST

from app_contact.models import User
from .models import HistoriqueRetentionSetting, HistoriqueUser
from .services import purge_old_history


def _get_client_ip(request) -> str:
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR") or ""


def _extract_session_key(details: str) -> str:
    """Extract session_key from HistoriqueUser.details (same logic as template filter)."""
    if not details:
        return ""
    marker = "session_key="
    if marker not in details:
        return ""
    value = details.split(marker, 1)[1]
    return value.split("|", 1)[0].strip()


def _extract_user_agent(details: str) -> str:
    """Extract user_agent from HistoriqueUser.details (same logic as template filter)."""
    if not details:
        return ""
    marker = "user_agent="
    if marker not in details:
        return ""
    value = details.split(marker, 1)[1]
    return value.split("|", 1)[0].strip()


@csrf_exempt
def log_client_url(request):
    """
    Endpoint appelé par JavaScript pour logger l'URL complète window.location.href
    (incluant le #fragment). Le serveur ne reçoit jamais le hash sinon.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    full_url = (payload.get("full_url") or "").strip()
    if not full_url:
        return JsonResponse({"ok": False, "error": "Missing full_url"}, status=400)

    # Assurer session_key même pour non connecté
    if not request.session.session_key:
        request.session.create()
    session_key = request.session.session_key or ""

    # User si connecté (sinon None)
    user = None
    user_session = request.session.get("user")
    if user_session and isinstance(user_session, dict) and user_session.get("id"):
        user = User.objects.filter(id=user_session["id"]).first()

    user_agent = (request.META.get("HTTP_USER_AGENT") or "").strip()
    ip = _get_client_ip(request)

    # ✅ Anti-spam : ne pas enregistrer 50 fois le même hash
    # Fenêtre courte (ex: 30 secondes)
    recent_window = timezone.now() - timedelta(seconds=30)

    already = (
        HistoriqueUser.objects.filter(
            action="READ",
            app_label="ui",
            model_name="page#fragment",
            url=full_url[:255],
            date_action__gte=recent_window.date(),  # approximation (date)
        )
        .order_by("-date_action", "-heure_action")
        .first()
    )

    # Comme on a date_action + heure_action séparés, on fait un check simple :
    # si on a un log aujourd’hui ET même session_key dans details, on ignore.
    if already and session_key and (f"session_key={session_key}" in (already.details or "")):
        return JsonResponse({"ok": True, "saved": False, "reason": "duplicate_recent"})

    details = (
        "URL complète côté navigateur (avec #fragment)"
        f" | session_key={session_key}"
        f" | user_agent={user_agent}"
    )

    try:
        HistoriqueUser.objects.create(
            user=user,  # None si non connecté
            action="READ",
            app_label="ui",
            model_name="page#fragment",
            object_id="",
            object_repr=full_url[:255],
            url=full_url[:255],
            method="POST",
            ip=ip[:64],
            details=details,
        )
    except Exception:
        return JsonResponse({"ok": True, "saved": False})

    return JsonResponse({"ok": True, "saved": True})


def _session_role(request) -> str:
    user = request.session.get("user") or {}
    if isinstance(user, dict):
        return str(user.get("role") or "")
    return ""


def _is_admin(request) -> bool:
    # In your DB, admin role is stored as lowercase "administrateur"
    return _session_role(request) == "administrateur"


@require_GET
def historique_fragment(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    retention_years = HistoriqueRetentionSetting.get_value()
    current_year = timezone.localdate().year
    years_kept = ", ".join(str(y) for y in range(current_year - retention_years + 1, current_year + 1))

    history_rows = (
        HistoriqueUser.objects.select_related("user")
        .all()
        .order_by("-date_action", "-heure_action")[:300]
    )

    return render(
        request,
        "app_historique/fragment_historique.html",
        {
            "history_rows": history_rows,
            "retention_years": retention_years,
            "current_year": current_year,
            "years_kept": years_kept,
        },
    )


@csrf_protect
@require_POST
def historique_settings(request):
    if not _is_admin(request):
        return JsonResponse({"ok": False, "message": "Accès refusé."}, status=403)

    if "btn_save_retention" in request.POST:
        try:
            value = int(request.POST.get("retention_years") or "2")
        except ValueError:
            value = 2
        value = max(1, min(10, value))

        setting, _ = HistoriqueRetentionSetting.objects.get_or_create(id=1)
        setting.retention_years = value
        setting.save()

        deleted = purge_old_history()
        return JsonResponse({"ok": True, "deleted": deleted, "message": f"Rétention enregistrée ({value} an(s))."})

    if "btn_purge_history" in request.POST:
        deleted = purge_old_history()
        return JsonResponse({"ok": True, "deleted": deleted, "message": "Purge effectuée."})

    return JsonResponse({"ok": False, "message": "Action inconnue."}, status=400)


def _rows_for_export():
    return (
        HistoriqueUser.objects.select_related("user")
        .all()
        .order_by("-date_action", "-heure_action")
    )


def _export_filename(ext: str) -> str:
    ts = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    return f"historique_utilisateur_{ts}.{ext}"


@require_GET
def export_historique_csv(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("csv")}"'

    writer = csv.writer(response)
    writer.writerow([
        "DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET",
        "SESSION", "NAVIGATEUR",
        "URL", "METHODE", "IP",
    ])

    for h in _rows_for_export():
        writer.writerow([
            h.date_action,
            h.heure_action,
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            _extract_session_key(h.details or ""),
            _extract_user_agent(h.details or ""),
            h.url,
            h.method,
            h.ip,
        ])
    return response


@require_GET
def export_historique_excel(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return HttpResponse(
            "Export Excel indisponible: installez la dépendance: pip install openpyxl",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"

    headers = [
        "DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET",
        "SESSION", "NAVIGATEUR",
        "URL", "METHODE", "IP",
    ]
    ws.append(headers)

    for h in _rows_for_export():
        ws.append([
            str(h.date_action),
            str(h.heure_action),
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            _extract_session_key(h.details or ""),
            _extract_user_agent(h.details or ""),
            h.url,
            h.method,
            h.ip,
        ])

    # Auto-width
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(70, len(header) + 18))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("xlsx")}"'
    return response


@require_GET
def export_historique_pdf(request):
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError:
        return HttpResponse(
            "Export PDF indisponible: installez la dépendance: pip install reportlab",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(15 * mm, height - 15 * mm, "Historique Utilisateur")
    c.setFont("Helvetica", 9)
    c.drawString(15 * mm, height - 21 * mm, f"Généré le {timezone.localtime().strftime('%d/%m/%Y %H:%M:%S')}")

    y = height - 30 * mm
    line_h = 6.5 * mm

    # ✅ Ajout: SESSION + NAVIGATEUR
    headers = ["DATE", "HEURE", "UTILISATEUR", "ACTION", "APP", "MODELE", "OBJET", "SESSION", "NAVIGATEUR", "URL"]

    # Largeurs: on limite "NAVIGATEUR" pour rester lisible (UA peut être long)
    fixed = [22, 18, 45, 18, 20, 26, 45, 25, 55]  # mm for first 9 cols (URL is remaining)
    cols = [w * mm for w in fixed]
    remaining = width - (15 * mm + sum(cols) + 15 * mm)
    cols.append(max(20 * mm, remaining))  # URL column

    x0 = 15 * mm

    def draw_row(values, y_pos, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 8.2)
        x = x0
        for idx, (val, w) in enumerate(zip(values, cols)):
            txt = "" if val is None else str(val)
            # Troncatures plus fortes pour UA/URL
            max_len = 60 if headers[idx] == "NAVIGATEUR" else (90 if headers[idx] == "URL" else 120)
            if len(txt) > max_len:
                txt = txt[: max_len - 3] + "..."
            c.drawString(x, y_pos, txt)
            x += w

    draw_row(headers, y, bold=True)
    y -= line_h
    c.line(x0, y + 2 * mm, width - 15 * mm, y + 2 * mm)

    for h in _rows_for_export():
        if y < 15 * mm:
            c.showPage()
            y = height - 15 * mm

        draw_row([
            h.date_action,
            h.heure_action,
            getattr(h.user, "email", "") if h.user else "",
            h.action,
            h.app_label,
            h.model_name,
            h.object_repr,
            _extract_session_key(h.details or ""),
            _extract_user_agent(h.details or ""),
            h.url,
        ], y, bold=False)
        y -= line_h

    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("pdf")}"'
    return response


@require_GET
def export_historique_json(request):
    """Export JSON (admin-only). Useful for backup/migration/analysis."""
    if not _is_admin(request):
        return HttpResponse("Accès refusé.", status=403)

    exported_at = timezone.localtime().isoformat()

    rows = []
    for h in _rows_for_export():
        details = h.details or ""
        rows.append({
            "date_action": str(h.date_action),
            "heure_action": str(h.heure_action),
            "user_email": (getattr(h.user, "email", "") if h.user else ""),
            "action": h.action,
            "app_label": h.app_label,
            "model_name": h.model_name,
            "object_id": h.object_id,
            "object_repr": h.object_repr,
            "session": _extract_session_key(details),
            "navigateur": _extract_user_agent(details),
            "url": h.url,
            "method": h.method,
            "ip": h.ip,
            "details": details,
        })

    payload = {
        "export_type": "historique_user",
        "version": 1,
        "exported_at": exported_at,
        "count": len(rows),
        "data": rows,
    }

    content = json.dumps(payload, ensure_ascii=False, indent=2)
    response = HttpResponse(content, content_type="application/json; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_export_filename("json")}"'
    return response
